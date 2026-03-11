"""
XLSX parser — reads spreadsheets and converts them to text.

Spreadsheets are structurally different from documents:
  - Data is in cells, not paragraphs
  - Multiple sheets may contain different document sections
  - Merged cells and empty rows are common

Strategy:
  - Iterate all sheets
  - Skip fully empty rows
  - Represent each sheet as a labeled section in the output text
"""
from pathlib import Path

import structlog

from .base import ParsedDocument, Parser

log = structlog.get_logger()


class XlsxParser(Parser):
    def parse(self, file_path: str | Path) -> ParsedDocument:
        import openpyxl

        path = self._validate_path(file_path)
        log.info("xlsx.parse.start", path=str(path))

        wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
        sections: list[str] = []
        tables: list[str] = []

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows = _extract_sheet_rows(ws)
            if not rows:
                continue

            # The sheet as a labeled text block
            section = f"[Sheet: {sheet_name}]\n" + "\n".join(rows)
            sections.append(section)
            tables.append(f"{sheet_name}:\n" + "\n".join(rows))

        wb.close()
        text = "\n\n".join(sections)
        log.info("xlsx.parse.done", path=str(path), sheets=len(sections))

        return ParsedDocument(
            text=text,
            source_path=str(path),
            format="xlsx",
            page_count=len(sections),
            tables=tables,
        )


def _extract_sheet_rows(ws) -> list[str]:
    rows: list[str] = []
    for row in ws.iter_rows(values_only=True):
        cells = [str(c).strip() if c is not None else "" for c in row]
        # Skip rows where every cell is empty
        if all(c == "" for c in cells):
            continue
        rows.append(" | ".join(cells))
    return rows
